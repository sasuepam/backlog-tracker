import openpyxl
import datetime

PATH = "/Users/sarahsuda/Documents/MSCbacklogtracker/Backlog Tracker - EINT MuleSoft.xlsx"

wb = openpyxl.load_workbook(PATH, data_only=False)
ws = wb["Backlog"]

# snapshot existing real IDs before touching anything (for post-fix diff)
pre_ids = set()
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 2).value  # Name column, real data rows only
    if v not in (None, ""):
        pre_ids.add(ws.cell(r, 1).value)

max_row = ws.max_row

for r in range(2, max_row + 1):
    ws.cell(r, 1).value = (
        f'=IF(B{r}="","","NEW-"&TEXT(SUMPRODUCT(MAX(IFERROR(VALUE(MID($A$2:$A{r-1},5,4)),0)))+1,"0000"))'
    )
    ws.cell(r, 9).value = (
        f'=IFERROR(IF(ISBLANK(INDEX(SprintPlanning[Sprint],MATCH($A{r},SprintPlanning[Requirement ID],0))),"",'
        f'INDEX(SprintPlanning[Sprint],MATCH($A{r},SprintPlanning[Requirement ID],0))),"")'
    )
    ws.cell(r, 11).value = (
        f'=IFERROR(IF(ISBLANK(INDEX(Release[Release],MATCH($A{r},Release[Requirement ID],0))),"",'
        f'INDEX(Release[Release],MATCH($A{r},Release[Requirement ID],0))),"")'
    )
    ws.cell(r, 15).value = f'=IF($N{r}="","","https://smartship.atlassian.net/browse/"&$N{r})'

# find first fully-blank row (no Name) to place the new item
target_row = None
for r in range(2, max_row + 1):
    if ws.cell(r, 2).value in (None, ""):
        target_row = r
        break
print("target_row", target_row)

ws.cell(target_row, 2).value = "Puerto Rico on US site Status Match (INT010.3)"
ws.cell(target_row, 3).value = "CR"
ws.cell(target_row, 4).value = "US"
ws.cell(target_row, 6).value = "NEW"
ws.cell(target_row, 10).value = 5
ws.cell(target_row, 12).value = datetime.datetime(2026, 8, 3)
ws.cell(target_row, 14).value = "MDTTPU-15251"

wb.save(PATH)
print("saved")
