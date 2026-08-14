import openpyxl
import datetime

PATH = "/Users/sarahsuda/Documents/MSCbacklogtracker/Backlog Tracker - EINT MuleSoft.xlsx"
wb = openpyxl.load_workbook(PATH, data_only=False)
bl = wb["Backlog"]
sp = wb["Sprint Planning"]

items = [
    {
        "row": 246,
        "name": "US Growth - B2CW Payment billing data/ US Release ACS Check - Part 1 (Adding the new fields to INT007, INT140, INT146)",
        "hle": 5,
        "jira": "MDTTPU-15252",
        "sprint": "Sprint 10",
        "id": "NEW-0408",
    },
    {
        "row": 247,
        "name": "US Growth - B2CW Payment billing data/ US Release ACS Check - Part 2 (Wiring the new fields to Datatrans in INT007, INT140, INT146)",
        "hle": 13,
        "jira": "MDTTPU-15253",
        "sprint": "Sprint 11",
        "id": "NEW-0409",
    },
    {
        "row": 248,
        "name": "US Growth - M4M Payment billing data/ US Release ACS Check (Adding the fields in M4M transaction init)",
        "hle": 8,
        "jira": "MDTTPU-15254",
        "sprint": "Sprint 11",
        "id": "NEW-0410",
    },
]

for it in items:
    r = it["row"]
    bl.cell(r, 2).value = it["name"]
    bl.cell(r, 4).value = "Growth"
    bl.cell(r, 6).value = "ON HOLD"
    bl.cell(r, 10).value = it["hle"]
    bl.cell(r, 12).value = datetime.datetime(2026, 8, 3)
    bl.cell(r, 14).value = it["jira"]

sp_start = 244
for i, it in enumerate(items):
    r = sp_start + i
    sp.cell(r, 1).value = it["id"]
    sp.cell(r, 2).value = (
        f'=IFERROR(IF(ISBLANK(INDEX(Backlog[Requirement Name],MATCH($A{r},Backlog[Requirement ID],0))),"",'
        f'INDEX(Backlog[Requirement Name],MATCH($A{r},Backlog[Requirement ID],0))),"")'
    )
    sp.cell(r, 3).value = (
        f'=IFERROR(IF(ISBLANK(INDEX(Backlog[Stream],MATCH($A{r},Backlog[Requirement ID],0))),"",'
        f'INDEX(Backlog[Stream],MATCH($A{r},Backlog[Requirement ID],0))),"")'
    )
    sp.cell(r, 4).value = (
        f'=IFERROR(IF(ISBLANK(INDEX(Backlog[Priority],MATCH($A{r},Backlog[Requirement ID],0))),"",'
        f'INDEX(Backlog[Priority],MATCH($A{r},Backlog[Requirement ID],0))),"")'
    )
    sp.cell(r, 5).value = (
        f'=IFERROR(IF(ISBLANK(INDEX(Backlog[Status],MATCH($A{r},Backlog[Requirement ID],0))),"",'
        f'INDEX(Backlog[Status],MATCH($A{r},Backlog[Requirement ID],0))),"")'
    )
    sp.cell(r, 6).value = it["sprint"]
    sp.cell(r, 7).value = (
        f'=IFERROR(IF(ISBLANK(INDEX(Backlog[HLE],MATCH($A{r},Backlog[Requirement ID],0))),"",'
        f'INDEX(Backlog[HLE],MATCH($A{r},Backlog[Requirement ID],0))),"")'
    )
    sp.cell(r, 8).value = f'=IFERROR(VALUE(SUBSTITUTE($F{r},"Sprint ","")),9999)'

# extend Sprint Planning table ref to cover new rows
tbl = sp.tables["SprintPlanning"]
old_ref = tbl.ref  # e.g. A1:H243
end_row = sp_start + len(items) - 1
tbl.ref = f"A1:H{end_row}"

wb.save(PATH)
print("Sprint Planning table ref:", old_ref, "->", tbl.ref)
print("saved")
