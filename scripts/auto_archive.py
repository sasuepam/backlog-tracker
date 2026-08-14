#!/usr/bin/env python3
"""
auto_archive.py — triggered by the Excel VBA event handler (or run manually).

Finds every Backlog row whose Status = DONE that isn't already in a Quarter tab,
then for each one:
  1. Appends a row to the correct Quarter tab (Sprint 1-4 → Q1, 5-8 → Q2, etc.)
  2. Freezes the Release tab's formula-pulled columns to literal values
  3. Deletes the row from Sprint Planning, RTM, API Design Planning
  4. Deletes the row from Backlog

Run manually:
    python3 /Users/sarahsuda/Documents/MSCbacklogtracker/scripts/auto_archive.py

Or triggered automatically by the Workbook_SheetChange VBA event.
"""
import openpyxl
import sys
import os

WORKBOOK = "/Users/sarahsuda/Documents/MSCbacklogtracker/workbook/Backlog Tracker - EINT MuleSoft.xlsx"

# Sprint → Quarter sheet name
_Q = {i: "Q1 Backlog Report 2026" for i in range(1, 5)}
_Q.update({i: "Q2 Backlog Report 2026" for i in range(5, 9)})
_Q.update({i: "Q3 Backlog Report 2026" for i in range(9, 13)})
_Q.update({i: "Q4 Backlog Report 2026" for i in range(13, 17)})

# Quarter sheet column order (must match what bulk archive wrote)
# A  B     C     D       E      F       G         H          I         J       K        L
# ID Name  Type  Stream  Owner  Status  JiraKey   JiraLink   Priority  Sprint  Release  HLE
Q_COLS = ["req_id","name","type","stream","owner","status","jira_key","jira_link","priority","sprint","release","hle"]


def sprint_number(sprint_val):
    """Extract integer from 'Sprint 5' → 5, or None."""
    if not sprint_val:
        return None
    import re
    m = re.search(r'\d+', str(sprint_val))
    return int(m.group()) if m else None


def quarter_sheet(sprint_val):
    n = sprint_number(sprint_val)
    return _Q.get(n) if n else None


def header_index(ws, keyword, required=False):
    """Return 0-based column index of the first header containing keyword (case-insensitive)."""
    for cell in ws[1]:
        if cell.value and keyword.lower() in str(cell.value).lower():
            return cell.column - 1
    if required:
        raise ValueError(f"No header containing '{keyword}' in sheet '{ws.title}'")
    return None


def find_row_by_id(ws, req_id, id_col=0):
    """Return 1-based row number where ws.cell(row, id_col+1) == req_id, or None."""
    for row in ws.iter_rows(min_row=2):
        if row[id_col].value and str(row[id_col].value).strip() == str(req_id).strip():
            return row[id_col].row
    return None


def already_archived_ids(wb):
    ids = set()
    for qname in _Q.values():
        if qname not in wb.sheetnames:
            continue
        ws_q = wb[qname]
        for row in ws_q.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ids.add(str(row[0]).strip())
    return ids


def get_release_for_id(ws_rel, req_id):
    id_col  = header_index(ws_rel, "req") or header_index(ws_rel, "id") or 0
    rel_col = header_index(ws_rel, "release")
    if rel_col is None:
        return None
    for row in ws_rel.iter_rows(min_row=2, values_only=True):
        if row[id_col] and str(row[id_col]).strip() == str(req_id).strip():
            return row[rel_col]
    return None


def get_sprint_for_id(ws_sp, req_id):
    id_col  = header_index(ws_sp, "req") or 0
    spr_col = header_index(ws_sp, "sprint")
    if spr_col is None:
        return None
    for row in ws_sp.iter_rows(min_row=2, values_only=True):
        if row[id_col] and str(row[id_col]).strip() == str(req_id).strip():
            return row[spr_col]
    return None


def append_to_quarter(wb, qname, data):
    ws_q = wb[qname]
    # Find next empty row
    last = 1
    for row in ws_q.iter_rows(min_row=2):
        if row[0].value:
            last = row[0].row
    next_row = last + 1
    for i, key in enumerate(Q_COLS):
        ws_q.cell(next_row, i + 1).value = data.get(key)


def freeze_release_row(ws_rel, req_id, data):
    id_col = header_index(ws_rel, "req") or header_index(ws_rel, "id") or 0
    row_num = find_row_by_id(ws_rel, req_id, id_col)
    if row_num is None:
        return
    hdr = [c.value for c in ws_rel[1]]
    freeze_map = {
        "name":     next((i for i,h in enumerate(hdr) if h and "name"     in h.lower()), None),
        "stream":   next((i for i,h in enumerate(hdr) if h and "stream"   in h.lower()), None),
        "sprint":   next((i for i,h in enumerate(hdr) if h and "sprint"   in h.lower()), None),
        "status":   next((i for i,h in enumerate(hdr) if h and "status"   in h.lower()), None),
        "priority": next((i for i,h in enumerate(hdr) if h and "priority" in h.lower()), None),
        "hle":      next((i for i,h in enumerate(hdr) if h and "hle"      in h.lower()), None),
    }
    for key, col_idx in freeze_map.items():
        if col_idx is not None:
            ws_rel.cell(row_num, col_idx + 1).value = data.get(key)


def delete_row_by_id(ws, req_id, id_col=None):
    if id_col is None:
        id_col = header_index(ws, "req") or header_index(ws, "id") or 0
    row_num = find_row_by_id(ws, req_id, id_col)
    if row_num:
        ws.delete_rows(row_num)


def archive_one(wb, req_id):
    ws_bl  = wb["Backlog"]
    ws_sp  = wb["Sprint Planning"]
    ws_rtm = wb["RTM"]
    ws_api = wb["API Design Planning"]
    ws_rel = wb["Release"]

    # --- Read Backlog row ---
    bl_id_col = 0  # Col A
    bl_row_num = find_row_by_id(ws_bl, req_id, bl_id_col)
    if bl_row_num is None:
        print(f"  SKIP {req_id}: not found in Backlog")
        return False

    def v(col):  # 1-based col → value from the Backlog row
        return ws_bl.cell(bl_row_num, col).value

    # For formula cells (Sprint=I=9, Jira Link=O=15) we need the stored formula text
    # but Sprint we get from Sprint Planning for reliability
    sprint_val = get_sprint_for_id(ws_sp, req_id)

    data = {
        "req_id":    v(1),   # A – Req ID
        "name":      v(2),   # B – Name
        "type":      v(3),   # C – Type
        "stream":    v(4),   # D – Stream
        "owner":     v(5),   # E – Stream Owner
        "status":    v(6),   # F – Status
        "priority":  v(7),   # G – Priority
        "hle":       v(10),  # J – HLE
        "jira_key":  v(14),  # N – Jira Key
        "jira_link":
            f"https://smartship.atlassian.net/browse/{v(14)}" if v(14) else "",
        "sprint":    sprint_val,
        "release":   get_release_for_id(ws_rel, req_id),
    }

    qname = quarter_sheet(sprint_val)
    if not qname:
        print(f"  SKIP {req_id}: cannot determine Quarter from sprint={sprint_val!r}")
        return False

    # 1 – add to quarter tab
    append_to_quarter(wb, qname, data)
    print(f"  {req_id} → {qname}")

    # 2 – freeze Release row
    freeze_release_row(ws_rel, req_id, data)

    # 3 – delete from Sprint Planning, RTM, API Design (before Backlog)
    delete_row_by_id(ws_sp,  req_id)
    delete_row_by_id(ws_rtm, req_id)
    delete_row_by_id(ws_api, req_id)

    # 4 – delete from Backlog last
    delete_row_by_id(ws_bl, req_id, bl_id_col)

    return True


def run(workbook_path=WORKBOOK):
    if not os.path.exists(workbook_path):
        print(f"ERROR: Workbook not found: {workbook_path}")
        sys.exit(1)

    print(f"Loading {os.path.basename(workbook_path)} ...")
    wb = openpyxl.load_workbook(workbook_path)

    archived = already_archived_ids(wb)

    # Find Done items in Backlog not yet archived
    ws_bl = wb["Backlog"]
    to_archive = []
    for row in ws_bl.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rid    = str(row[0]).strip()
        status = str(row[5]).strip().upper() if row[5] else ""
        if status == "DONE" and rid not in archived:
            to_archive.append(rid)

    if not to_archive:
        print("Nothing to archive — no new Done items found.")
        return

    print(f"Archiving {len(to_archive)} item(s):")
    count = sum(archive_one(wb, rid) for rid in to_archive)

    if count:
        print(f"Saving ...")
        wb.save(workbook_path)
        print(f"Done. {count} item(s) archived.")
    else:
        print("No items were archived (all were skipped).")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else WORKBOOK
    run(path)
