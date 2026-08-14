import openpyxl
import re

PATH = "/Users/sarahsuda/Documents/MSCbacklogtracker/Backlog Tracker - EINT MuleSoft.xlsx"
wb = openpyxl.load_workbook(PATH, data_only=False)

SHEETS = ["Sprint Planning", "RTM", "Release", "API Design Planning"]

ref_re = re.compile(r'(\$[A-Z]{1,2})(\d+)')

total_fixed = 0
for name in SHEETS:
    ws = wb[name]
    fixed = 0
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                new_v = ref_re.sub(lambda m: f"{m.group(1)}{r}", v)
                if new_v != v:
                    cell.value = new_v
                    fixed += 1
    print(name, "cells fixed:", fixed)
    total_fixed += fixed

wb.save(PATH)
print("total fixed", total_fixed)
